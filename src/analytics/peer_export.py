import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

conn = sqlite3.connect("db/nifty100.db")

peer_groups = pd.read_excel(
    "data/raw/peer_groups.xlsx"
)

percentiles = pd.read_sql(
    """
    SELECT *
    FROM peer_percentiles
    """,
    conn
)

financial = pd.read_sql(
    """
    SELECT *
    FROM financial_ratios
    """,
    conn
)

companies = pd.read_sql(
    """
    SELECT
        id AS company_id,
        company_name
    FROM companies
    """,
    conn
)

peer_groups = pd.read_sql(
    """
    SELECT
        company_id,
        peer_group_name,
        is_benchmark
    FROM peer_groups
    """,
    conn
)

financial["year_num"] = (
    financial["year"]
    .str.extract(r"(\d{4})")[0]
    .astype(int)
)

financial = (
    financial
    .sort_values("year_num")
    .drop_duplicates(
        subset="company_id",
        keep="last"
    )
)

master = (
    financial
    .merge(companies, on="company_id", how="left")
    .merge(peer_groups, on="company_id", how="left")
)

percentile_table = (
    percentiles
    .pivot_table(
        index="company_id",
        columns="metric",
        values="percentile_rank"
    )
    .reset_index()
)

percentile_table.columns = [
    c if c == "company_id"
    else f"{c}_pctile"
    for c in percentile_table.columns
]

master = master.merge(
    percentile_table,
    on="company_id",
    how="left"
)

print(master.shape)

print(
    master[
        [
            "company_id",
            "company_name",
            "peer_group_name"
        ]
    ].head()
)

output_file = "output/peer_comparison.xlsx"

with pd.ExcelWriter(
    output_file,
    engine="openpyxl"
) as writer:

    for peer in sorted(
        master["peer_group_name"]
        .dropna()
        .unique()
    ):

        sheet = (
            master[
                master["peer_group_name"] == peer
            ]
            .sort_values(
                "composite_quality_score",
                ascending=False
            )
        )

        sheet.to_excel(
            writer,
            sheet_name=peer[:31],
            index=False
        )

print("Excel Created!")
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(output_file)

green_fill = PatternFill(
    fill_type="solid",
    start_color="C6EFCE"
)

yellow_fill = PatternFill(
    fill_type="solid",
    start_color="FFF2CC"
)

red_fill = PatternFill(
    fill_type="solid",
    start_color="F4CCCC"
)

gold_fill = PatternFill(
    fill_type="solid",
    start_color="FFD966"
)

green = PatternFill(
    fill_type="solid",
    start_color="C6EFCE"
)

yellow = PatternFill(
    fill_type="solid",
    start_color="FFEB9C"
)

red = PatternFill(
    fill_type="solid",
    start_color="FFC7CE"
)

gold = PatternFill(
    fill_type="solid",
    start_color="FFD966"
)

percentile_columns = [
    "asset_turnover_pctile",
    "debt_to_equity_pctile",
    "eps_cagr_5yr_pctile",
    "free_cash_flow_cr_pctile",
    "interest_coverage_pctile",
    "net_profit_margin_pct_pctile",
    "operating_profit_margin_pct_pctile",
    "pat_cagr_5yr_pctile",
    "return_on_equity_pct_pctile",
    "revenue_cagr_5yr_pctile",
]

for sheet in wb.worksheets:

    headers = [
        cell.value
        for cell in sheet[1]
    ]

    # ---------------------------------
    # Color Percentile Columns
    # ---------------------------------

    for col in percentile_columns:

        if col not in headers:
            continue

        col_index = headers.index(col) + 1

        for row in range(2, sheet.max_row + 1):

            cell = sheet.cell(
                row=row,
                column=col_index
            )

            if cell.value is None:
                continue

            if cell.value >= 75:
                cell.fill = green_fill

            elif cell.value <= 25:
                cell.fill = red_fill

            else:
                cell.fill = yellow_fill

        if "is_benchmark" in headers:

            benchmark_col = (
                headers.index("is_benchmark") + 1
            )

        for row in range(2, sheet.max_row + 1):

            value = sheet.cell(
                row=row,
                column=benchmark_col
            ).value

            if value in [1, True, "True"]:

                for col in range(
                    1,
                    sheet.max_column + 1
                ):
                    sheet.cell(
                        row=row,
                        column=col
                    ).fill = gold_fill

        numeric_columns = [
        "return_on_equity_pct",
        "operating_profit_margin_pct",
        "net_profit_margin_pct",
        "debt_to_equity",
        "free_cash_flow_cr",
        "pat_cagr_5yr",
        "revenue_cagr_5yr",
        "eps_cagr_5yr",
        "interest_coverage",
        "asset_turnover",
    ]

    median_row = sheet.max_row + 2

    sheet.cell(
        row=median_row,
        column=1
    ).value = "Peer Median"

    for metric in numeric_columns:

        if metric not in headers:
            continue

        col = headers.index(metric) + 1

        values = []

        for row in range(2, sheet.max_row + 1):

            value = sheet.cell(
                row=row,
                column=col
            ).value

            if isinstance(value, (int, float)):
                values.append(value)

        if values:

            sheet.cell(
                row=median_row,
                column=col
            ).value = round(
                pd.Series(values).median(),
                2
            )

wb.save(output_file)

print("\nPeer comparison report created successfully!")
